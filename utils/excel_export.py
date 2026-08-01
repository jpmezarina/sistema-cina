"""
Exportación del Record Clínico a Excel - Sistema CINA
Genera un libro con una hoja "Record" que resume, por estudiante,
los procedimientos validados por tratamiento y las notas finales,
replicando el formato de la planilla original en Excel.
"""
import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from config import Config, BASE_DIR
from models import Usuario, Procedimiento, NotaEstudiante

HEADER_FILL = PatternFill(start_color='1A2634', end_color='1A2634', fill_type='solid')
HEADER_FONT = Font(color='FFFFFF', bold=True, size=10)
TITLE_FONT = Font(bold=True, size=14, color='1A2634')
SUBTITLE_FONT = Font(italic=True, size=10, color='666666')
THIN_BORDER = Border(
    left=Side(style='thin', color='D0D0D0'),
    right=Side(style='thin', color='D0D0D0'),
    top=Side(style='thin', color='D0D0D0'),
    bottom=Side(style='thin', color='D0D0D0'),
)
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT = Alignment(horizontal='left', vertical='center')

VERDE = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
AMARILLO = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
ROJO = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')


def _contar_validados(estudiante_id, codigo_tratamiento, procedimientos_por_estudiante):
    procs = procedimientos_por_estudiante.get(estudiante_id, [])
    return sum(1 for p in procs if p.tratamiento_codigo == codigo_tratamiento and p.estado == 'validado')


def exportar_record_excel():
    """
    Construye el archivo Excel con el record de todos los estudiantes
    y lo guarda en disco, devolviendo la ruta del archivo generado.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = 'Record'

    tratamientos = Config.TRATAMIENTOS
    estudiantes = Usuario.query.filter_by(rol='estudiante').order_by(Usuario.apellidos).all()

    # Pre-cargar procedimientos y notas para evitar N+1 queries
    todos_procedimientos = Procedimiento.query.all()
    procedimientos_por_estudiante = {}
    for p in todos_procedimientos:
        procedimientos_por_estudiante.setdefault(p.estudiante_id, []).append(p)

    notas_por_estudiante = {n.estudiante_id: n for n in NotaEstudiante.query.all()}

    # ---------- Encabezado del documento ----------
    total_cols = 3 + len(tratamientos) + 5  # N, Estudiante, Cod. Acceso + tratamientos + Nota Clin/Caso/Actitud/Trabajo/Final
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    ws.cell(row=1, column=1, value='CINA 2026 - RECORD AUTOMATIZADO DE CLÍNICA INTEGRAL').font = TITLE_FONT
    ws.cell(row=1, column=1).alignment = CENTER

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
    ws.cell(row=2, column=1, value=f'Generado el {datetime.now().strftime("%d/%m/%Y %H:%M")}').font = SUBTITLE_FONT
    ws.cell(row=2, column=1).alignment = CENTER

    # ---------- Encabezados de tabla (fila 4) ----------
    header_row = 4
    headers = ['N°', 'Código', 'Apellidos y Nombres']
    for t in tratamientos:
        headers.append(t['codigo'])
    headers += ['Nota Clínica (70%)', 'Caso Clínico (10%)', 'Actitudinal (10%)', 'Trabajo Acad. (10%)', 'NOTA FINAL']

    for col_idx, htext in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=htext)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    # ---------- Filas de datos ----------
    row = header_row + 1
    for idx, est in enumerate(estudiantes, start=1):
        col = 1
        ws.cell(row=row, column=col, value=idx).alignment = CENTER; col += 1
        ws.cell(row=row, column=col, value=est.codigo_acceso).alignment = CENTER; col += 1
        ws.cell(row=row, column=col, value=est.nombre_completo).alignment = LEFT; col += 1

        for t in tratamientos:
            count = _contar_validados(est.id, t['codigo'], procedimientos_por_estudiante)
            requerido = t['requerido']
            cell = ws.cell(row=row, column=col, value=f"{count}/{requerido}")
            cell.alignment = CENTER
            cell.fill = VERDE if count >= requerido else (AMARILLO if count > 0 else ROJO)
            col += 1

        notas = notas_por_estudiante.get(est.id)
        nota_clinica = round(notas.nota_clinica, 2) if notas else 0
        nota_caso = notas.nota_caso_clinico if notas and notas.nota_caso_clinico is not None else 0
        nota_actitud = notas.nota_actitudinal if notas and notas.nota_actitudinal is not None else 0
        nota_trabajo = notas.nota_trabajo_academico if notas and notas.nota_trabajo_academico is not None else 0
        nota_final = round(notas.nota_final, 2) if notas else 0

        for val in [nota_clinica, nota_caso, nota_actitud, nota_trabajo]:
            ws.cell(row=row, column=col, value=val).alignment = CENTER
            col += 1

        final_cell = ws.cell(row=row, column=col, value=nota_final)
        final_cell.alignment = CENTER
        final_cell.font = Font(bold=True)
        final_cell.fill = VERDE if nota_final >= 13 else (ROJO if nota_final < 10.5 else AMARILLO)

        for c in range(1, total_cols + 1):
            ws.cell(row=row, column=c).border = THIN_BORDER

        row += 1

    # ---------- Anchos de columna ----------
    ws.column_dimensions[get_column_letter(1)].width = 5
    ws.column_dimensions[get_column_letter(2)].width = 14
    ws.column_dimensions[get_column_letter(3)].width = 32
    for i in range(4, 4 + len(tratamientos)):
        ws.column_dimensions[get_column_letter(i)].width = 10
    for i in range(4 + len(tratamientos), total_cols + 1):
        ws.column_dimensions[get_column_letter(i)].width = 14

    ws.freeze_panes = ws.cell(row=header_row + 1, column=4)

    # ---------- Guardar archivo ----------
    export_dir = os.path.join(BASE_DIR, 'static', 'exports')
    os.makedirs(export_dir, exist_ok=True)
    filename = f'CINA_2026_I_RECORD_AUTOMATIZADO_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    filepath = os.path.join(export_dir, filename)
    wb.save(filepath)

    return filepath
